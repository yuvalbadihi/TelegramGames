#from requests import Request, Session
import time
import json
import pprint
import os
from openpyxl import load_workbook #Workbook
import tgMethods
import telethon
import gameClasses
from threading import Thread


LOCAL_CONFIGURATION_FILE = "configuration.json"

with open(LOCAL_CONFIGURATION_FILE) as local_config_file:
    local_config = json.load(local_config_file)


token = local_config["Details"]["token"]
bot_username = local_config["Details"]["bot_username"]
badi_bot = tgMethods.BotHandler(token)
bot_name = local_config["Details"]["bot_name"]
group_owner_username = local_config["Details"]["group_owner_username"]
my_tg_id = local_config["Details"]["my_tg_id"]
droid_core = local_config["Details"]["droid_core"]
currency = local_config["Details"]["currency"]
currency_symbol = "ϟ"
groupforum = local_config["Details"]["groupforum"]
my_fm = local_config["Details"]["my_fm"]
database_file = local_config["Details"]["database_file"]

wb = load_workbook(database_file)
ws = wb.active

Games_Lst = []
Slots_Games_Lst = []

def check_if_reply(current_update):
    try:
        text = current_update['message']['text']
        if (('profile' in text) and ('reply_to_message' not in current_update['message'])):
            if  current_update['message']['from']['id'] == current_update['message']['chat']['id']:
                badi_bot.send_message(current_update['message']['from']['id'],"❌ In order to use this command please reply to someones message")
                return
            badi_bot.send_message(current_update['message']['from']['id'], "❌ In order to use this command please reply to someones message")
            badi_bot.send_message(current_update['message']['chat']['id'], "❌ In order to use this command please reply to someones message")
    except:
        print("Reply Error")

def view_profile(current_update):
    user_id = current_update['message']['reply_to_message']['from']['id']
    firstname = current_update['message']['reply_to_message']['from']['first_name']
    player1 = gameClasses.Player(user_id, firstname)
    message_text = f"👨🏽‍💼 {player1.firstname} 𝗣𝗿𝗼𝗳𝗶𝗹𝗲 \n\nID: FM{user_id}\nFame: {player1.get_value_or_zero(wb,ws,'I')}"
    button = {"inline_keyboard": [[{"text": '📜 History', "callback_data": 'get_history'},{"text": '🛒 Shop', "callback_data": 'get_someones_shop'}],[{"text": '❤ Fame', "callback_data": 'fame_to_someone'},{"text": '💔 Defame', "callback_data": 'defame_to_someone'}]]}
    badi_bot.send_inline_callback_button(current_update['message']['from']['id'], message_text, button)
    badi_bot.send_inline_callback_button(current_update['message']['chat']['id'], message_text, button)
    return True

def get_user_id_from_profile(profile_text):
    new_lst = profile_text.split("ID: ")
    new_lst = new_lst[1].split('\n')
    return new_lst[0][2:]

def get_firstname_from_profile(profile_text):
    new_lst = profile_text.split(" 𝗣𝗿𝗼𝗳𝗶𝗹𝗲")
    return new_lst[0][4:]



def shopping(user_id_request,user_id_shop,firstname):

    player = gameClasses.Player(user_id_shop, firstname)
    shop = player.get_a_value(wb,ws,'K')
    badi_bot.send_message(user_id_request,f"🏪 {player.firstname} 𝗦𝘁𝗼𝗿𝗲\n\n{shop}")

    return True

def fame_defame(chat_id,famer_user_id,famer_firstname,famed_user_id,famed_firstname,change,my_username):
    if change == 1:
        stat = 'fame'
        emoji = '➕'
    else:
        stat = 'defame'
        emoji = '➖'

    famer = gameClasses.Player(famer_user_id, famer_firstname)
    famed = gameClasses.Player(famed_user_id, famed_firstname)
    if (famed.user_id == famer.user_id):
        badi_bot.send_message(famer.user_id, f'❌ You can not {stat} your self')
        return False
    min = famer.fame_someone(wb,ws,famed,change,my_username)

    if min == True:
        badi_bot.send_message(famer.user_id,f'{emoji} You {stat}d {famed.firstname}')
        badi_bot.send_message(famed.user_id, f'{emoji} You got {stat}d by {famer.firstname}')
       # badi_bot.send_message(chat_id, f'{emoji} {famed.firstname} been {stat}d by {famer.firstname}')
    else:
        badi_bot.send_message(famer.user_id, f'❌ You will be able to {stat} again in {min}minutes')



def business(current_update):
    if 'message' in current_update:
        user_id = current_update['message']['from']['id']
        firstname = current_update['message']['from']['first_name']
        chat_id = current_update['message']['chat']['id']
    else:
        user_id = current_update['callback_query']['from']['id']
        firstname = current_update['callback_query']['from']['first_name']
        chat_id = current_update['callback_query']['message']['chat']['id']

    player1 = gameClasses.Player(user_id, firstname)

    message_text = f"👨🏽‍💼 {player1.firstname} 𝗣𝗿𝗼𝗳𝗶𝗹𝗲 \n\nID: FM{user_id}\nFame: {player1.get_value_or_zero(wb,ws,'I')}"

    button = {"inline_keyboard": [[{"text": '🏷️ Set Store', "callback_data": 'set_store'},
                                   {"text": '🛒 My Store', "callback_data": 'get_someones_shop'}],[{"text": '🏦 Free Markets', "callback_data": 'get_free_markets'},{"text": '📜 Fame History', "callback_data": 'get_history'}]]}
    badi_bot.send_inline_callback_button(chat_id, message_text, button)


def getFMText(top_10, num):
    text = f''
    text += f'𝗧𝗼𝗽 {num} 𝗙𝗿𝗲𝗲 𝗠𝗮𝗿𝗸𝗲𝘁𝘀\n\n'
    for i in range(len(top_10)):
        text += top_10[len(top_10) - i - 1].addy
        text += f"   |   {top_10[len(top_10) - i - 1].rank} {currency_symbol}\n"
    return text


def bubbleSortFM(arr):
    n = len(arr)
    for i in range(n - 1):
        for j in range(0, n - i - 1):
            if int(arr[j].rank) > int(arr[j + 1].rank):
                arr[j], arr[j + 1] = arr[j + 1], arr[j]


def freemarket(chat_id):
    top_list_size = 3
    top_10 = []

    j = 2
    for i in ws:
        fm_instance = gameClasses.a_fm_instance(str(ws['E' + str(j)].value), str(ws['F' + str(j)].value))

        if fm_instance.addy == 'None':
            j += 1
            continue

        if len(top_10) < top_list_size:
            top_10.append(fm_instance)

        else:
            mini = fm_instance
            mini_index = -1
            for n in range(top_list_size):
                if int(top_10[n].rank) < int(mini.rank):
                    mini_index = n
                    mini = top_10[n]

            if mini_index != -1:
                top_10[mini_index] = fm_instance

        j += 1

    bubbleSortFM(top_10)
    message_text = getFMText(top_10, top_list_size)
    print(message_text)
    button = {"inline_keyboard": [
        [{"text": '🏪 Set FM', "callback_data": 'set_fm'}, {"text": '⚡ Boost FM', "callback_data": 'boost_fm'}]]}
    x = badi_bot.send_inline_callback_button(chat_id, message_text, button)
    print(x)


def start_slots_func(game):
    jackpot = 0
    inline_msg_id = game.inline_message_id
    x = telethon.utils.resolve_inline_message_id(inline_msg_id)
    game.message_id = x[0]
    game.chat_id = telethon.utils.get_peer_id(x[1], add_mark=True)
    f = badi_bot.send_message(game.chat_id, f'🎰 {game.player1_firstname} Spins')
    if str(f) != '<Response [200]>':
        print("Good!")
    # a_new_game = multiprocessing.Process(target=start_new_game,args=[game])
    # a_new_game.start()
    try:
        player1 = gameClasses.Player(game.player1_id, game.player1_firstname)
        player1_chips = player1.chips(wb, ws)
        if player1_chips < game.amount:
            return False
        player1.reduce_chips(wb, ws, game.amount)

        winning_value = getRandomRes(game.chat_id, '🎰')

        winning_value = int(winning_value)
        if winning_value >= 0 and winning_value <= 10:  # [1..10]
            jackpot = int(float(game.amount) * 0.2)

        elif winning_value >= 11 and winning_value <= 20:  # [11..20]
            jackpot = int(float(game.amount) * 0.3)

        elif winning_value >= 21 and winning_value <= 30:  # [21..30]
            jackpot = int(float(game.amount) * 0.7)

        elif winning_value >= 31 and winning_value <= 40:
            jackpot = int(float(game.amount) * 1)

        elif winning_value >= 41 and winning_value <= 50:
            jackpot = int(float(game.amount) * 1.5)

        elif winning_value >= 51 and winning_value <= 60:
            jackpot = int(float(game.amount) * 2)

        elif winning_value >= 61 and winning_value <= 63:
            jackpot = int(float(game.amount) * 3)

        elif winning_value == 64:
            jackpot = int(float(game.amount) * 15)

        player1.increase_chips(wb, ws, jackpot)
        wb.save(database_file)
        time.sleep(1)
        if jackpot >= game.amount:
            emoji = '🍀 '
        else:
            emoji = ' ☘️ '

        badi_bot.send_message(game.chat_id, emoji + player1.firstname + ' won ' + str(jackpot) + f' {currency}!')
        time.sleep(1)
        game.limit += 1
        if game.limit == 3:
            button_slots = {"inline_keyboard": [[{"text": '🛠️ Repair', "callback_data": 'repair_now'}]]}
            button_slots = json.dumps(button_slots)
            badi_bot.edit_message_inline_id(game.inline_message_id, "The slots machine broke down", button_slots)
            game.lock = 1

    except:
        print('error')
        # badi_bot.send_message(game.chat_id,'❌ Error, please restart the game')
    return True


def all_chips(first_chat_id):
    slots_amount = 0  # ws['M' + str(2)].value
    owners = ''
    total = 0
    j = 2
    for i in ws:
        if str(ws['A' + str(j)].value) == 'None' or ws['C' + str(j)].value == 0:
            j += 1
            continue
        owners += str(ws['A' + str(j)].value) + ' own ' + str(ws['C' + str(j)].value) + f' {currency}.\n'
        if str(ws['C' + str(j)].value).isdigit():
            total += ws['C' + str(j)].value

        j += 1
    badi_bot.send_message(first_chat_id,
                          owners + '\n________\nTotal: ' + str(total) + ' + ' + str(slots_amount) + ' = ' + str(
                              total + slots_amount))
    j = 0


def show_wallet(current_update):
    user_id = current_update['message']['from']['id']
    firstname = current_update['message']['from']['first_name']
    chat_id = current_update['message']['chat']['id']

    player1 = gameClasses.Player(user_id, firstname)
    playercoins = player1.chips(wb, ws)
    message_text = f'{firstname} {currency} 𝘄𝗮𝗹𝗹𝗲𝘁 𝗯𝗮𝗹𝗮𝗻𝗰𝗲:\n\n{playercoins} ©️'
    badi_bot.send_message(chat_id, message_text)


def addchips(current_update):
    text = current_update['message']['text']
    text_lst = text.split(' ')
    amount = int(text_lst[2])
    user_id = int(text_lst[1])
    player1 = gameClasses.Player(user_id, 'idc')
    player1.increase_chips(wb, ws, amount)
    badi_bot.send_message(user_id, f'✅ You received {amount} {currency}.')
    badi_bot.send_message(droid_core, f'✅ You sent {amount} {currency}.')


def start_dices_func(game):
    try:
        winner = None
        Games_Lst.remove(game)
        inline_msg_id = game.inline_message_id
        x = telethon.utils.resolve_inline_message_id(inline_msg_id)
        game.message_id = x[0]
        game.chat_id = telethon.utils.get_peer_id(x[1], add_mark=True)
        print(game.chat_id)
        f = badi_bot.send_message(game.chat_id, '🎲 Game started!\nAnd may the odds be ever in your favor.')
        print(f)
        if str(f) != '<Response [200]>':
            print("Error")
        # a_new_game = multiprocessing.Process(target=start_new_game,args=[game])
        # a_new_game.start()
        player1 = gameClasses.Player(game.player1_id, game.player1_firstname)
        player2 = gameClasses.Player(game.player2_id, game.player2_firstname)
        player1_chips = player1.chips(wb, ws)
        player2_chips = player2.chips(wb, ws)
        jackpot = 0
        if player1_chips < game.amount or player2_chips < game.amount:
            return False
        jackpot = game.amount * 2
        player1.reduce_chips(wb, ws, game.amount)
        player2.reduce_chips(wb, ws, game.amount)
        player1_wins = 0
        player2_wins = 0
        while player1_wins == player2_wins:
            badi_bot.send_message(game.chat_id, player1.firstname + ' roll: ')
            num_player1 = roll(game.chat_id)
            badi_bot.send_message(game.chat_id, player2.firstname + ' roll: ')
            num_player2 = roll(game.chat_id)

            if num_player1 > num_player2:
                player1_wins += 1
                winner = player1

            elif num_player2 > num_player1:
                player2_wins += 1
                winner = player2

        winner.increase_chips(wb, ws, jackpot)
        wb.save(database_file)
        badi_bot.send_message(game.chat_id, '🍀 ' + winner.firstname + ' won ' + str(game.amount) + f' {currency}!')
    except:
        print('error')
        # badi_bot.send_message(game.chat_id,'❌ Error, please restart the game')
    return True


def awnser_to_inline_dice(inline_query_id):
    # WORK EXAMPLE
    button_slots = {"inline_keyboard": [[{"text": '📍 Spin', "callback_data": 'spin_now'}],
                                        [{"text": f'1 {currency_symbol}', "callback_data": 'add-1-chip'},
                                         {"text": f'2 {currency_symbol}', "callback_data": 'add-2-chip'},
                                         {"text": f'4 {currency_symbol}', "callback_data": 'add-4-chip'},
                                         {"text": f'8 {currency_symbol}', "callback_data": 'add-8-chip'},
                                         {"text": f'16 {currency_symbol}', "callback_data": 'add-16-chip'}]]}
    button_dices = {"inline_keyboard": [[{"text": '🕹️ Join Match', "callback_data": 'join_dices'},
                                         {"text": '🎲 Start', "callback_data": 'start_dices'}],
                                        [{"text": f'1 {currency_symbol}', "callback_data": 'add_1_chip'},
                                         {"text": f'2 {currency_symbol}', "callback_data": 'add_2_chip'},
                                         {"text": f'4 {currency_symbol}', "callback_data": 'add_4_chip'},
                                         {"text": f'8 {currency_symbol}', "callback_data": 'add_8_chip'},
                                         {"text": f'16 {currency_symbol}', "callback_data": 'add_16_chip'}]]}
    dices = {'description': 'start a new dices match',
             'thumb_url': 'https://cdn.pixabay.com/photo/2012/04/05/01/24/dice-25637_1280.png', 'title': 'Dice Game',
             'type': 'article', 'id': 'request_to_start_dices',
             'input_message_content': {'message_text': f'Hey, I would like to play dices with you on \n0 {currency}.'},
             'reply_markup': button_dices}
    slots = {'description': 'Slots gameplay',
             'thumb_url': 'https://cdn.pixabay.com/photo/2013/07/13/12/37/slot-machine-159972_1280.png',
             'title': 'Slots', 'type': 'article', 'id': 'request_to_start_slots', 'input_message_content': {
            'message_text': '🎰 If you’re looking for a thrill, then you better take chances on playing a 3-reel slot game!\n\n𝗠𝘂𝗹𝘁𝗶𝗽𝗹𝗶𝗲𝗿: 0'},
             'reply_markup': button_slots}
    results1 = [dices, slots]
    results1 = json.dumps(results1)
    badi_bot.answer_inline_query(inline_query_id, results1)


def pm(chat_id, text):
    badi_bot.send_message(int(chat_id), text)


def check_amount(amount, player1):
    # return int = 0 incase of error
    # else retrun chips amount
    playerchips = player1.chips(wb, ws)
    numstest = amount.isdigit()
    if numstest == False:
        return 0
    amount = int(amount)
    if amount > playerchips or amount <= 0:
        return 0
    return amount


def amoun_test_max_min(min_chips, max_chips, player_amount):
    if player_amount <= max_chips and player_amount >= min_chips:
        return True
    return 0


def withdraw(user_id, chat_id, firstname, message_id):
    player1 = gameClasses.Player(user_id, firstname)
    with_addy = player1.withdrawal_addy(wb, ws)
    text = f'Your current withdrawal {currency} address is:\n\n{with_addy}\n\n Tap Continue to select an amount to send.'
    button = {"inline_keyboard": [[{"text": '📝 Set new withdrawal address', "callback_data": 'new_address'}],
                                  [{"text": '⏩ Continue', "callback_data": 'select_amount'}]]}

    badi_bot.edit_message(chat_id, message_id, text)
    badi_bot.edit_inline(chat_id, message_id, button)

def change_values_for_updates(player1,chat_id,message_id):
    cooldown = player1.get_mine_val(wb, ws, 'G') * gameClasses.basic_cooldown
    pick_axe = int(player1.get_mine_val(wb, ws, 'H') * gameClasses.basic_pickaxe)

    hoursleft = int(cooldown // 3600)
    minsleft = int((cooldown % 3600) // 60)
    secleft = int(cooldown % 60)

    cooldown_price = int(player1.get_improve_price(wb, ws, 'G'))
    pickaxe_price = int(player1.get_improve_price(wb, ws, 'H'))
    text = f'{player1.firstname} 𝗦𝘁𝗮𝘁𝘀\n\n⏳ Mining cooldown: {hoursleft}H {minsleft}M {secleft}S\n⛏ Mining rate: {pick_axe}'
    button = {"inline_keyboard": [
        [{"text": f'➕ 5% Cooldown decrease ({cooldown_price}{currency_symbol})', "callback_data": 'cooldown_drop'}], [
            {"text": f'➕ 5% Rate increase ({pickaxe_price}{currency_symbol})', "callback_data": 'pick_axe_boost'}],
        ]}

    badi_bot.edit_message(chat_id, message_id, text)
    badi_bot.edit_inline(chat_id, message_id,button)


def bold(str1):
    str2 = ''
    str3 = '𝟬𝟭𝟮𝟯𝟰𝟱𝟲𝟳𝟴𝟵'
    for i in str1:
        str2 += str3[int(i)]
    return str2


def wallet_button(current_update):
    if 'message' in current_update:
        user_id = current_update['message']['from']['id']
        firstname = current_update['message']['from']['first_name']
        chat_id = current_update['message']['chat']['id']
    else:
        user_id = current_update['callback_query']['from']['id']
        firstname = current_update['callback_query']['from']['first_name']
        chat_id = current_update['callback_query']['message']['chat']['id']

    player1 = gameClasses.Player(user_id, firstname)
    playercoins = player1.chips(wb, ws)
    message_text = f'𝗬𝗼𝘂𝗿 {currency} 𝘄𝗮𝗹𝗹𝗲𝘁 𝗯𝗮𝗹𝗮𝗻𝗰𝗲:\n\n{playercoins} {currency_symbol}'

    button = {"inline_keyboard": [[{"text": '📥 Deposit', "callback_data": 'deposit_chips'},
                                   {"text": '📤 Withdraw', "callback_data": 'withdraw_chips'}],
                                  [{"text": '📊️ Stats', "callback_data": 'get_stats'},{"text": '⛏️ Mine', "callback_data": 'mine_chips'}]]}
    badi_bot.send_inline_callback_button(chat_id, message_text, button)


def private_keyboard(current_update):
    chat_id = current_update['message']['chat']['id']
    text = current_update['message']['text']

    if text == '/start':
        keyboard = {"resize_keyboard": True, 'one_time_keyboard': False,
                    "keyboard": [[{"text": "💰 Wallet"}, {"text": "🎲 Play"}],
                                 [{"text": "💡 Support"}, {"text": "💼 Business"}]]}
        keyboard = json.dumps(keyboard)
        badi_bot.open_keyboard(chat_id, 'Successfully completed.', keyboard)


def getRandomRes(first_chat_id, emoji):
    while True:
        r_dict = None
        r = badi_bot.send_dice(first_chat_id, emoji)
        # time.sleep(1)
        r_dict = r.json()
        if 'result' in r_dict:
            break
    time.sleep(1)
    winning_value = r_dict['result']['dice']['value']
    winning_value = str(winning_value)
    print(winning_value)
    return (winning_value)


def roll(first_chat_id):
    winning_num = None
    winning_value = getRandomRes(first_chat_id, '🎲')

    if winning_value == '1':
        winning_num = '1️⃣'
    if winning_value == '2':
        winning_num = '2️⃣'
    if winning_value == '3':
        winning_num = '3️⃣'
    if winning_value == '4':
        winning_num = '4️⃣'
    if winning_value == '5':
        winning_num = '5️⃣'
    if winning_value == '6':
        winning_num = '6️⃣'
    badi_bot.send_message(first_chat_id, '💃 Number: ' + winning_num)
    time.sleep(2)
    return winning_num


def restart():
    print("Restarting...")
    os.system("python telegramGames.py")
    time.sleep(10)
    return True


def main_menu_tggames():
    new_offset = 0
    first_update_id = 0
    timeout_restart = time.time() + 3600  # Auto Restart
    timeout_spin = time.time()
    time_mine = time.time()

    print('Leggo \n')
    while True:

        if time.time() > timeout_restart:
            timeout_restart = time.time() + 3600
            restart()

        all_updates = badi_bot.get_updates(new_offset)
        if len(all_updates) > 0:
            for current_update in all_updates:
                current_update_dumped = json.dumps(current_update)
                print('_____________________\n')
                pprint.pprint(json.loads(current_update_dumped))
                print('_____________________\n')
                first_update_id = current_update['update_id']

                if 'chosen_inline_result' in current_update:
                    if current_update['chosen_inline_result']['result_id'] == 'request_to_start_dices':
                        Games_Lst_item = gameClasses.NewGame(None, None,
                                                             current_update['chosen_inline_result']['from']['id'], 0,
                                                             current_update['chosen_inline_result'][
                                                                 'inline_message_id'],
                                                             current_update['chosen_inline_result']['from'][
                                                                 'first_name'], None, None)
                        Games_Lst.append(Games_Lst_item)

                    if current_update['chosen_inline_result']['result_id'] == 'request_to_start_slots':
                        Games_Lst_item = gameClasses.NewSlotsGame(None, None,
                                                                  current_update['chosen_inline_result']['from']['id'],
                                                                  0, current_update['chosen_inline_result'][
                                                                      'inline_message_id'],
                                                                  current_update['chosen_inline_result']['from'][
                                                                      'first_name'])
                        Slots_Games_Lst.append(Games_Lst_item)

                if 'message' in current_update:
                    if 'chat' in current_update['message']:
                        check_if_reply(current_update)
                        chat_id = current_update['message']['chat']['id']
                        if current_update['message']['chat']['type'] == 'private':
                            if 'text' in current_update['message']:
                                private_keyboard(current_update)
                    if 'text' in current_update['message']:
                        if '/allchips' in current_update['message']['text'] and int(
                                current_update['message']['from']['id']) == int(my_tg_id):
                            all_chips(chat_id)

                        if '/wallet' in current_update['message']['text']:
                            show_wallet(current_update)

                        if current_update['message']['text'] == "💡 Support" and current_update['message']['chat']['type'] == 'private':
                            badi_bot.force_reply(chat_id,
                                                 f'🆘 In order to get support please check out our forum {groupforum}.\n if you still need help reply to this message.\n\n⬇️ Enter Support Ticket ⬇️')

                        if current_update['message']['text'] == "💼 Business" and current_update['message']['chat']['type'] == 'private':
                            business(current_update)

                        if current_update['message']['text'] == "💰 Wallet" and current_update['message']['chat']['type'] == 'private':
                            wallet_button(current_update)

                        if current_update['message']['text'] == '🎲 Play' and current_update['message']['chat']['type'] == 'private':
                            badi_bot.send_inline_button(chat_id, '🎲',
                                                        f'💬 Choose conversation to play on\n\n⚠️Make sure that {bot_name} found inside of the chat you would like to play on.',
                                                        '👁️‍🗨️ Select')

                        if '/addchips' in current_update['message']['text'] and int(
                                current_update['message']['from']['id']) == int(my_tg_id):
                            addchips(current_update)

                        if '/pm' in current_update['message']['text'] and str(
                                current_update['message']['from']['id']) == str(my_tg_id):
                            lst1 = (current_update['message']['text']).split(' ')
                            user_id_to_pm = lst1[1]
                            lst1.pop(0)
                            lst1.pop(0)
                            text_to_pm = ''
                            for i in lst1:
                                text_to_pm += i
                                text_to_pm += ' '

                            print(lst1)
                            pm(user_id_to_pm, text_to_pm)

                    if 'reply_to_message' in current_update['message']:
                        try:
                            new_addy = current_update['message']['text']
                            user_id = current_update['message']['chat']['id']
                            firstname = current_update['message']['chat']['first_name']

                            if current_update['message']['reply_to_message']['text'] == '✏ Insert your stores link and a short description:':
                                player1 = gameClasses.Player(user_id, firstname)
                                qtest = player1.set_new_shop(new_addy, wb, ws)
                                if qtest == True:
                                    badi_bot.send_message(chat_id, '✅ New shop established successfully')
                                else:
                                    badi_bot.send_message(chat_id, '❌ Error, please enter less than 200 characters')

                            if current_update['message']['reply_to_message']['text'] == '✏️ Insert new address:':

                                player1 = gameClasses.Player(user_id, firstname)
                                qtest = player1.set_withdrawal_addy(new_addy, wb, ws)
                                if qtest == True:
                                    badi_bot.send_message(chat_id, '✅ New address inserted successfully')
                                else:
                                    badi_bot.send_message(chat_id, '❌ Error')

                            if current_update['message']['reply_to_message'][
                                'text'] == f'✏️ Insert your free market tag\n(e.g. {my_fm})':

                                player1 = gameClasses.Player(user_id, firstname)
                                qtest = player1.set_fm_tag(new_addy, wb, ws)
                                if qtest == True:
                                    badi_bot.send_message(chat_id, '✅ New address inserted successfully')
                                else:
                                    badi_bot.send_message(chat_id, '❌ Error')

                            if current_update['message']['reply_to_message'][
                                'text'] == f'✏️ Insert the amount of {currency} you wish to invest to promote your Free Market:':
                                boost = new_addy
                                player1 = gameClasses.Player(user_id, firstname)
                                qtest = player1.set_fm_boost(boost, wb, ws)
                                if qtest == True:
                                    badi_bot.send_message(chat_id,
                                                          f'✅ {boost} {currency_symbol} added!\n{player1.getFM(wb, ws)} rank is {player1.getFMRank(wb, ws)}')
                                    print(
                                        f'✅ {boost} {currency} added!\n{player1.getFM(wb, ws)} rank is {player1.getFMRank(wb, ws)}{currency_symbol}')
                                else:
                                    badi_bot.send_message(chat_id, '❌ Error')

                        except:
                            print("An exception occurred 1")

                    if 'reply_to_message' in current_update['message']:
                        try:

                            if 'text' in current_update['message']['reply_to_message']:
                                if '/profile' in current_update['message'][
                                    'text'] :
                                    view_profile(current_update)

                                if current_update['message']['reply_to_message'][
                                    'text'] == f'🆘 In order to get support please check out our forum {groupforum}.\n if you still need help reply to this message.\n\n⬇️ Enter Support Ticket ⬇️':
                                    firstname = current_update['message']['chat']['first_name']
                                    ticket = current_update['message']['text']
                                    user_id = current_update['message']['chat']['id']
                                    badi_bot.send_message(droid_core, str(user_id))
                                    badi_bot.send_message(droid_core, str(user_id) + '\nTicket 👇')
                                    badi_bot.send_message(droid_core, ticket)

                                if current_update['message']['reply_to_message'][
                                    'text'] == f'💳 Insert the amount of {currency} you wish to withdraw:':

                                    amount = current_update['message']['text']
                                    user_id = current_update['message']['chat']['id']
                                    firstname = current_update['message']['chat']['first_name']
                                    player_send = gameClasses.Player(user_id, firstname)
                                    # LOCK
                                    amount_test = check_amount(amount, player_send)
                                    player_recive_id = str(player_send.withdrawal_addy(wb, ws))[2:]
                                    if player_recive_id.isdigit():
                                        player_recive_id = int(player_recive_id)
                                    player_recive = gameClasses.Player(player_recive_id, 'doesnt matter')

                                    if amount_test != 0:
                                        max_chips = 10000
                                        min_chips = 10
                                        amount_test_max_min = amoun_test_max_min(min_chips, max_chips, amount_test)
                                        if current_update['message']['reply_to_message'][
                                            'text'] == f'💳 Insert the amount of {currency} you wish to withdraw:':
                                            min_chips = 1
                                            amount_test_max_min = amoun_test_max_min(min_chips, max_chips, amount_test)

                                        if amount_test_max_min == True:
                                            player_send.reduce_chips(wb, ws, amount_test)
                                            player_recive.increase_chips(wb, ws, amount_test)
                                            badi_bot.send_message(chat_id,
                                                                  f'✅ {amount_test} {currency} sent successfully')
                                            badi_bot.send_message(player_recive.user_id,
                                                                  f'✅ {amount_test} {currency} recived successfully from {firstname}')


                                        else:
                                            badi_bot.send_message(chat_id,
                                                                  f'❌ Error please select higher amount than {min_chips}')
                                    else:
                                        badi_bot.send_message(chat_id, '❌ Error')
                        except:

                            print("An exception occurred 2")

                if 'callback_query' in current_update:
                    user_id = current_update['callback_query']['from']['id']
                    firstname = current_update['callback_query']['from']['first_name']
                    callback = current_update['callback_query']['data']
                    call_back_id = current_update['callback_query']['id']

                    if 'add_' in callback:
                        for game in Games_Lst:
                            if game.player1_id == user_id and game.inline_message_id == \
                                    current_update['callback_query']['inline_message_id']:
                                x = callback.split('_')
                                x = int(x[1])
                                game.change_chips(x, call_back_id, wb, ws, currency_symbol, bot_username, badi_bot,
                                                  currency)
                                print(game.amount)

                            elif game.player1_id != user_id and game.inline_message_id == \
                                    current_update['callback_query']['inline_message_id']:
                                badi_bot.awnser_call_back_alert(call_back_id, '❌ You are not the creator of this game.')

                    elif 'add-' in callback:
                        for game in Slots_Games_Lst:
                            if game.player1_id == user_id and game.inline_message_id == \
                                    current_update['callback_query']['inline_message_id']:
                                x = callback.split('-')
                                x = int(x[1])
                                game.change_chips(x, call_back_id, wb, ws, currency_symbol, bot_username, badi_bot,
                                                  currency)
                                print(game.amount)

                            elif game.player1_id != user_id and game.inline_message_id == \
                                    current_update['callback_query']['inline_message_id']:
                                badi_bot.awnser_call_back_alert(call_back_id, '❌ You are not the creator of this game.')

                    if callback == 'join_dices':
                        player2 = gameClasses.Player(user_id, firstname)
                        player2_chips = player2.chips(wb, ws)
                        for game in Games_Lst:
                            if player2.user_id != game.player1_id and game.inline_message_id == \
                                    current_update['callback_query'][
                                        'inline_message_id'] and player2_chips >= game.amount and game.amount > 0:
                                game.player2_id = player2.user_id
                                game.player2_firstname = player2.firstname
                                button_dices = {"resize_inline_keyboard": True, "inline_keyboard": [
                                    [{"text": '🕹️ Join Match', "callback_data": 'join_dices'},
                                     {"text": '🎲 Start', "callback_data": 'start_dices'}],
                                    [{"text": f'1 {currency_symbol}', "callback_data": 'add_1_chip'},
                                     {"text": f'2 {currency_symbol}', "callback_data": 'add_2_chip'},
                                     {"text": f'4 {currency_symbol}', "callback_data": 'add_4_chip'},
                                     {"text": f'8 {currency_symbol}', "callback_data": 'add_8_chip'},
                                     {"text": f'16 {currency_symbol}', "callback_data": 'add_16_chip'}]]}
                                button_dices = json.dumps(button_dices)
                                badi_bot.edit_message_inline_id(game.inline_message_id,
                                                                f'Hey, I would like to play dices with you on \n{game.amount} {currency}.\n{game.player1_firstname} vs {game.player2_firstname}',
                                                                button_dices)
                            elif player2_chips < game.amount and player2.user_id != game.player1_id and game.inline_message_id == \
                                    current_update['callback_query']['inline_message_id']:
                                badi_bot.awnser_call_back_alert(call_back_id,
                                                                f'❌ Unfortunately you do not have enough {currency} to complete this operation,\nMake sure to mine {currency} through {bot_username} before trying again.')
                            elif player2.user_id == game.player1_id and game.inline_message_id == \
                                    current_update['callback_query'][
                                        'inline_message_id'] and player2_chips >= game.amount:
                                badi_bot.awnser_call_back_alert(call_back_id, "❌ You can't play with yourself")

                            elif player2.user_id != game.player1_id and game.inline_message_id == \
                                    current_update['callback_query'][
                                        'inline_message_id'] and player2_chips >= game.amount and game.amount == 0:
                                badi_bot.awnser_call_back_alert(call_back_id, "❌ You can't join none {currency} game")

                    if callback == 'start_dices':
                        for game in Games_Lst:
                            if game.player1_id == user_id and game.player2_id != None and game.inline_message_id == \
                                    current_update['callback_query']['inline_message_id']:
                                try:
                                    chat_id = game.chat_id
                                    # start_dices_func(game)
                                    t = Thread(target=start_dices_func, args=(game,))
                                    t.start()

                                except:
                                    print("An exception occurred 3")
                                    badi_bot.awnser_call_back_alert(call_back_id,
                                                                    f'❌ Failed!\nPlease add {bot_username} to this chat before trying again.\nThanks!')
                            elif game.player1_id != user_id and game.player2_id != None and game.inline_message_id == \
                                    current_update['callback_query']['inline_message_id']:
                                badi_bot.awnser_call_back_alert(call_back_id, '❌ You are not the creator of this game.')
                            elif game.player2_id == None and game.player1_id == user_id and game.inline_message_id == \
                                    current_update['callback_query']['inline_message_id']:
                                badi_bot.awnser_call_back_alert(call_back_id,
                                                                "❌ Please be patient and wait for another player to join this game.")

                    if callback == 'repair_now':  # useless atm
                        for game in Slots_Games_Lst:
                            if game.inline_message_id == current_update['callback_query']['inline_message_id']:
                                button_slots = {"inline_keyboard": [[{"text": '📍 Spin', "callback_data": 'spin_now'}],
                                                                    [{"text": f'1 {currency_symbol}',
                                                                      "callback_data": 'add-1-chip'},
                                                                     {"text": f'2 {currency_symbol}',
                                                                      "callback_data": 'add-2-chip'},
                                                                     {"text": f'4 {currency_symbol}',
                                                                      "callback_data": 'add-4-chip'},
                                                                     {"text": f'8 {currency_symbol}',
                                                                      "callback_data": 'add-8-chip'},
                                                                     {"text": f'16 {currency_symbol}',
                                                                      "callback_data": 'add-16-chip'}]]}
                                button_slots = json.dumps(button_slots)
                                badi_bot.edit_message_inline_id(game.inline_message_id,
                                                                f"🎰 If you’re looking for a thrill, then you better take chances on playing a 3-reel slot game!\n\n𝗠𝘂𝗹𝘁𝗶𝗽𝗹𝗶𝗲𝗿: {game.amount}",
                                                                button_slots)
                                game.lock = 0

                    if callback == 'spin_now':
                        # if  time.time() > timeout_spin:
                        min_spin = 5
                        player1 = gameClasses.Player(user_id, firstname)
                        player_chips = player1.chips(wb, ws)
                        for game in Slots_Games_Lst:
                            if game.inline_message_id == current_update['callback_query'][
                                'inline_message_id'] and player_chips >= game.amount and game.amount >= min_spin and game.lock == 0:
                                try:
                                    game.player1_id = user_id
                                    game.player1_firstname = firstname
                                    chat_id = game.chat_id
                                    t = Thread(target=start_slots_func, args=(game,))
                                    t.start()
                                # spin_done = start_slots_func(game)
                                # if spin_done == True:
                                # timeout_spin = time.time() + 15
                                except:
                                    print("An exception occurred 4")
                                    badi_bot.awnser_call_back_alert(call_back_id,
                                                                    f'❌ Failed!\nPlease add {bot_username} to this chat before trying again.\nThanks!')

                            elif game.inline_message_id == current_update['callback_query'][
                                'inline_message_id'] and player_chips >= game.amount and game.amount < min_spin:
                                badi_bot.awnser_call_back_alert(call_back_id,
                                                                f"❌ You can't spin on less than {min_spin} {currency}")

                    # elif time.time() < timeout_spin:
                    #   badi_bot.awnser_call_back_alert_black(call_back_id,'⌛ Cooldown Remaining Time: '+str(int(timeout_spin-time.time()))+' seconds.')

                    if 'message' in current_update['callback_query']:
                        chat_id = current_update['callback_query']['message']['chat']['id']
                        message_id = current_update['callback_query']['message']['message_id']

                    if current_update['callback_query']['data'] == 'deposit_chips':
                        badi_bot.send_message(chat_id, f'Your personal {currency} address:')
                        player1 = gameClasses.Player(user_id, firstname)
                        player_coins = 'FM' + str(player1.addy(wb, ws))
                        badi_bot.send_message(chat_id, player_coins)

                    elif current_update['callback_query']['data'] == 'withdraw_chips':
                        withdraw(user_id, chat_id, firstname, message_id)

                    if current_update['callback_query']['data'] == 'new_address':
                        badi_bot.force_reply(chat_id, '✏️ Insert new address:')

                    if current_update['callback_query']['data'] == 'select_amount':
                        badi_bot.force_reply(chat_id, f'💳 Insert the amount of {currency} you wish to withdraw:')

                    if current_update['callback_query']['data'] == 'mine_chips':
                        player1 = gameClasses.Player(user_id, firstname)
                        timeout_mine = player1.check_mine_time(wb, ws,'D')
                        if time.time() > timeout_mine:
                            added_amount = player1.mine(time.time(), wb, ws)
                            badi_bot.send_message(user_id,
                                                  f'✅ Mining completed (+{added_amount}{currency_symbol})')


                        elif time.time() < timeout_mine:
                            timeleft = int(timeout_mine - time.time())
                            hoursleft = timeleft // 3600
                            minsleft = timeleft // 60
                            secleft = timeleft % 60
                            badi_bot.awnser_call_back_alert_black(call_back_id, '⌛ Cooldown Remaining Time: ' + str(
                                hoursleft) + ':' + str(minsleft) + ':' + str(secleft))

                    if current_update['callback_query']['data'] == 'set_fm':
                        badi_bot.force_reply(chat_id, f'✏️ Insert your free market tag\n(e.g. {my_fm})')

                    if current_update['callback_query']['data'] == 'boost_fm':
                        badi_bot.force_reply(chat_id,
                                             f'✏️ Insert the amount of {currency} you wish to invest to promote your Free Market:')


                    if current_update['callback_query']['data'] == 'get_stats':
                        player1 = gameClasses.Player(user_id, firstname)
                        cooldown = player1.get_mine_val(wb,ws,'G') * gameClasses.basic_cooldown
                        pick_axe = int(player1.get_mine_val(wb, ws, 'H') * gameClasses.basic_pickaxe)

                        hoursleft = int(cooldown // 3600)
                        minsleft = int((cooldown%3600) // 60)
                        secleft = int(cooldown % 60)

                        cooldown_price = int(player1.get_improve_price(wb, ws,'G'))
                        pickaxe_price = int(player1.get_improve_price(wb,ws,'H'))

                        message_text = f'{player1.firstname} 𝗦𝘁𝗮𝘁𝘀\n\n⏳ Mining cooldown: {hoursleft}H {minsleft}M {secleft}S\n⛏ Mining rate: {pick_axe}'
                        button = {"inline_keyboard": [[{"text": f'➕ 5% Cooldown decrease ({cooldown_price}{currency_symbol})', "callback_data": 'cooldown_drop'}],[
                                                       {"text": f'➕ 5% Rate increase ({pickaxe_price}{currency_symbol})', "callback_data": 'pick_axe_boost'}],
                                                     ]}
                        badi_bot.send_inline_callback_button(chat_id, message_text, button)

                    if current_update['callback_query']['data'] == 'pick_axe_boost' or current_update['callback_query']['data'] == 'cooldown_drop':
                        data = current_update['callback_query']['data']
                        player1 = gameClasses.Player(user_id, firstname)
                        cooldown_price = player1.get_improve_price(wb, ws, 'G')
                        pickaxe_price = player1.get_improve_price(wb,ws,'H')


                        if data == 'pick_axe_boost':
                            price = player1.get_improve_price(wb, ws, 'H')
                            if price <= player1.chips(wb,ws):
                                player1.reduce_chips(wb,ws,price)
                                player1.improve_item(wb,ws,'H',1.05)
                                change_values_for_updates(player1, chat_id, message_id)
                            else:
                                badi_bot.send_message(chat_id,'❌ Error')

                        elif data == 'cooldown_drop':
                            price = player1.get_improve_price(wb, ws, 'G')
                            if price <= player1.chips(wb,ws):
                                player1.reduce_chips(wb,ws,price)
                                player1.improve_item(wb,ws,'G',0.95)
                                change_values_for_updates(player1, chat_id, message_id)
                            else:
                                badi_bot.send_message(chat_id,'❌ Error')


                        cooldown = player1.get_mine_val(wb,ws,'G') * gameClasses.basic_cooldown
                        pick_axe = int(player1.get_mine_val(wb, ws, 'H') * gameClasses.basic_pickaxe)
                        hoursleft = int(cooldown // 3600)
                        minsleft = int((cooldown%3600) // 60)
                        secleft = int(cooldown % 60)
                        message_text = f'{player1.firstname} 𝗦𝘁𝗮𝘁𝘀\n\n⏳ Mining cooldown: {hoursleft}H {minsleft}M {secleft}S\n⛏ Mining rate: {pick_axe}'
                        button = {"inline_keyboard": [[{"text": f'➕ 5% Cooldown decrease ({cooldown_price}{currency_symbol})', "callback_data": 'cooldown_drop'}],[
                                                       {"text": f'➕ 5% Rate increase ({pickaxe_price}{currency_symbol})', "callback_data": 'pick_axe_boost'}],
                                                     ]}
                    if current_update['callback_query']['data'] == 'get_free_markets':
                        freemarket(chat_id)

                    if current_update['callback_query']['data'] == 'set_store':
                        badi_bot.force_reply(chat_id,  '✏ Insert your stores link and a short description:')

                    if current_update['callback_query']['data'] == 'get_someones_shop':
                        user_id_request = current_update['callback_query']['from']['id']
                        text = current_update['callback_query']['message']['text']
                        user_id_shop = int(get_user_id_from_profile(text))
                        firstname = get_firstname_from_profile(text)
                        shopping(user_id_request,user_id_shop,firstname)

                    if current_update['callback_query']['data'] == 'fame_to_someone' or current_update['callback_query']['data'] == 'defame_to_someone':
                        famer_user_id = current_update['callback_query']['from']['id']
                        famer_firstname = current_update['callback_query']['from']['first_name']
                        text = current_update['callback_query']['message']['text']
                        chat_id = current_update['callback_query']['message']['chat']['id']
                        famed_user_id = int(get_user_id_from_profile(text))
                        famed_firstname = get_firstname_from_profile(text)
                        if 'username' in current_update['callback_query']['from']:
                            my_username = '@'+current_update['callback_query']['from']['username']
                        else:
                            my_username = 'Unknow'


                        if current_update['callback_query']['data'] == 'fame_to_someone':
                             fame_defame(chat_id, famer_user_id, famer_firstname, famed_user_id, famed_firstname,1,my_username)
                        else:
                             fame_defame(chat_id, famer_user_id, famer_firstname, famed_user_id, famed_firstname,-1,my_username)

                    if current_update['callback_query']['data'] == 'get_history':
                        text = current_update['callback_query']['message']['text']
                        request_by_id = famer_user_id = current_update['callback_query']['from']['id']
                        user_id = int(get_user_id_from_profile(text))
                        firstname = get_firstname_from_profile(text)
                        player1 = gameClasses.Player(user_id, firstname)
                        history = player1.show_history(wb,ws)
                        badi_bot.send_message(request_by_id,history)





                if 'inline_query' in current_update:
                    if current_update['inline_query']['query'] == '🎲':
                        inline_q_id = current_update['inline_query']['id']
                        awnser_to_inline_dice(inline_q_id)

                new_offset = first_update_id + 1


if __name__ == '__main__':
    try:
        main_menu_tggames()


    except KeyboardInterrupt:
        print('\nrestart \  stop')

        x = input('>')
        if x == 'restart':
            restart()

            exit()
        elif x == 'stop':
            exit()

        else:
            print('failed')






