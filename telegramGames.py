from requests import Request, Session
import time
import json
import pprint
import os
from openpyxl import Workbook , load_workbook
import tgmethods
import telethon


class Player:
    def __init__(self, user_id,firstname):
            self.user_id = user_id
            self.firstname = firstname
      
        
    def line(self,wb,ws):  
      #Lock_Wallets(wallets_lock)
      j = 1
      line = 0
      for i in ws:
         if self.user_id == ws['A' + str(j)].value:
             line = j
         j+=1

      if line == 0:
           ws.append([self.user_id])
           line = j
           ws['A' + str(j)].value = self.user_id
           ws['C' + str(j)].value = 0
        
      wb.save('wallets.xlsx')
      #UnLock_Wallets(wallets_lock)
      return line      
       
    def withdrawal_addy(self,wb,ws):
        coins = self.line(wb,ws)
        #Lock_Wallets(wallets_lock) 
        with_addy = ws['B' + str(coins)].value
        #UnLock_Wallets(wallets_lock)
        return with_addy   
    
    def set_withdrawal_addy(self,addy,wb,ws):
        coins = self.line(wb,ws)
        #Lock_Wallets(wallets_lock)
        if len(addy) < 50:    
            ws['B' + str(coins)].value = addy
            #UnLock_Wallets(wallets_lock)
            return True 
        return False
        
                
    
    def addy(self,wb,ws):
        coins = self.line(wb,ws)
        #Lock_Wallets(wallets_lock) 
        addy = ws['A' + str(coins)].value
        #UnLock_Wallets(wallets_lock)
        return addy    
        
    def chips(self,wb,ws):
        coins = self.line(wb,ws)
        #Lock_Wallets(wallets_lock) 
        user_coins = ws['C' + str(coins)].value
        #UnLock_Wallets(wallets_lock)
        return user_coins
    
    def increase_chips(self,wb,ws,amount):
        #LOCK
        line = self.line(wb, ws)
        ws['C' + str(line)].value += amount
        wb.save('wallets.xlsx')
        #UNLOCK
        return True
    
    def reduce_chips(self,wb,ws,amount):
        #LOCK
        line = self.line(wb, ws)
        ws['C' + str(line)].value -= amount
        wb.save('wallets.xlsx')
        #UNLOCK
        return True
   
    
    def check_mine_time(self):
        line = self.line(wb, ws)
        if ws['D' + str(line)].value == None:
            return time.time()-1
        return ws['D' + str(line)].value
        
    def mine(self,current_time):
        line = self.line(wb, ws)
        ws['D' + str(line)].value = time.time() + 3600
        ws['C' + str(line)].value += 2
        badi_bot.send_message(self.user_id,f'✅ Mining was successfully completed \n(+2 {currency_symbol})')
        wb.save('wallets.xlsx')

        
        
class NewGame:
    def __init__(self,message_id,chat_id,player1_id,amount,inline_message_id,player1_firstname,player2_firstname,player2_id):
            self.player1_id = player1_id
            self.player1_firstname = player1_firstname
            
            self.player2_id = player2_id
            self.player2_firstname = player2_firstname          
            
            self.message_id=message_id
            self.chat_id = chat_id
            self.amount = amount
            self.inline_message_id = inline_message_id
            
            
    def change_chips(self,new_amount,call_back_id):
        #LOCK
        player1 = Player(self.player1_id,self.player1_firstname)
        player1_chips = player1.chips(wb, ws)    
        if new_amount >= 0 and new_amount+self.amount <= player1_chips:
            button_dices = {"resize_inline_keyboard": True ,"inline_keyboard" : [[{"text" :'🕹️ Join Match', "callback_data" : 'join_dices'},{"text" :'🎲 Start', "callback_data" : 'start_dices'}],[{"text" :f'1 {currency_symbol}', "callback_data" : 'add-1-chip'},{"text" :f'2 {currency_symbol}', "callback_data" : 'add-2-chip'},{"text" :f'4 {currency_symbol}', "callback_data" : 'add-4-chip'},{"text" :f'8 {currency_symbol}', "callback_data" : 'add-8-chip'},{"text" :f'16 {currency_symbol}', "callback_data" : 'add-16-chip'}]]} 
            button_dices = json.dumps(button_dices)     
            self.amount += new_amount
            self.player2_id = None
            self.player2_firstname = None
            badi_bot.edit_message_inline_id(self.inline_message_id,f'Hey, I would like to play dices with you on \n{self.amount} {currency}.',button_dices)
     
        elif new_amount+self.amount > player1_chips:
                badi_bot.awnser_call_back_alert(call_back_id,f'❌ Unfortunately you do not have enough {currency} to complete this operation,\nMake sure to buy {currency} through {bot_username} before trying again.')
        #UNLOCK
        return True  
           
            
    #def play
    #def wait for another player 
    
class NewSlotsGame:
    def __init__(self,message_id,chat_id,player1_id,amount,inline_message_id,player1_firstname):
            self.player1_id = player1_id
            self.player1_firstname = player1_firstname           
            self.message_id=message_id
            self.chat_id = chat_id
            self.amount = amount
            self.inline_message_id = inline_message_id
            self.limit = 0
            self.lock = 0 
            
            
    def change_chips(self,new_amount,call_back_id):
        #LOCK
        player1 = Player(self.player1_id,self.player1_firstname)
        player1_chips = player1.chips(wb, ws)    
        if new_amount >= 0 and new_amount+self.amount <= player1_chips:
            button_slots = {"inline_keyboard" : [[{"text" :'📍 Spin', "callback_data" : 'spin_now'}],[{"text" :f'1 {currency_symbol}', "callback_data" : 'add-1-chip'},{"text" :f'2 {currency_symbol}', "callback_data" : 'add-2-chip'},{"text" :f'4 {currency_symbol}', "callback_data" : 'add-4-chip'},{"text" :f'8 {currency_symbol}', "callback_data" : 'add-8-chip'},{"text" :f'16 {currency_symbol}', "callback_data" : 'add-16-chip'}]]} 
            button_slots = json.dumps(button_slots)     
            self.amount += new_amount
            badi_bot.edit_message_inline_id(self.inline_message_id,f"🎰 If you’re looking for a thrill, then you better take chances on playing a 3-reel slot game!\n\n𝗠𝘂𝗹𝘁𝗶𝗽𝗹𝗶𝗲𝗿: {self.amount}",button_slots)
     
        elif new_amount+self.amount > player1_chips:
                badi_bot.awnser_call_back_alert(call_back_id,f'❌ Unfortunately you do not have enough {currency} to complete this operation,\nMake sure to mine {currency} through {bot_username} before trying again.')
        #UNLOCK
        return True  
           
            
    #def play
    #def wait for another player 
    
        

            
     
    
 ################################################################################
#########################     𝐆𝐋𝐎𝐁𝐀𝐋 𝐈𝐍𝐅𝐎    ##################################
################################################################################


token = 'xxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxx' # your bot token (through @botfather)
bot_username ='@xxxxxxxxx' #add your bot username (through @botfather)
badi_bot = tgmethods.BotHandler(token) #Your bot's name
bot_name = 'xxxxxxx'
group_owner_username = '@xxxxxxx' #add your account username (found in settings)
my_tg_id = 'XXXXXXXXX' #your telegram id (you can activate the bot once to get it on console)
droid_core = '-xxxxxxxxx'#Create new group > make it public > add the bot > make the bot admin > insent the group id inside of the string 
#you will get all of the support ticket to there and you can also use /addchips or /pm from there.
freemarket = 't.me/xxxxxxx' #FM group
currency = 'xxxxxx' #Pick ur own coin name
currency_symbol = 'x' #.. 
groupforum = '@xxxxxx' #you can enter a forum chat to assist the users

wallets_lock = True #doesnt work yet but do not remove it
wb = load_workbook('wallets.xlsx') #make sure to put the excel file in the same folder
ws = wb.active 

'''

BEFORE USING MAKE SURE TO ADD
1. wallets.xlsx
2. tgmethods.py
3. fill the bot_id in tgmethods.py
4. fill the info above

'''


################################################################################
################################################################################
################################################################################

Games_Lst =[]
Slots_Games_Lst =[]

def start_slots_func(game):
    jackpot = 0 
    inline_msg_id = game.inline_message_id
    x = telethon.utils.resolve_inline_message_id(inline_msg_id)
    game.message_id = x[0] 
    game.chat_id = telethon.utils.get_peer_id(x[1],add_mark=True)
    print(game.chat_id)
    f = badi_bot.send_message(game.chat_id,f'🎰 {game.player1_firstname} Spins')
    print(f)
    if str(f) != '<Response [200]>':
        badi_bot.send_message()
   # a_new_game = multiprocessing.Process(target=start_new_game,args=[game])
   # a_new_game.start()
    try:
        player1 = Player(game.player1_id,game.player1_firstname)
        player1_chips = player1.chips(wb, ws)
        if player1_chips < game.amount:
            return False
        player1.reduce_chips(wb, ws, game.amount)
        
        r = badi_bot.send_dice(game.chat_id,'🎰')
        r_dict = r.json()
        while 'result' not in r_dict:
            time.sleep(1)
        winning_value = r_dict['result']['dice']['value']
        winning_value = str(winning_value)
        print(winning_value)
        winning_value = int(winning_value)
        if winning_value >= 0 and winning_value <= 10: #[1..10]
            jackpot = int(float(game.amount)*0.1)
            
        elif winning_value >= 11 and winning_value <= 20: #[11..20]
            jackpot = int(float(game.amount)*0.2)
            
        elif winning_value >= 21 and winning_value <= 30: #[21..30]
            jackpot = int(float(game.amount)*0.5)
        
        elif winning_value >= 31 and winning_value <= 40: 
            jackpot = int(float(game.amount)*0.7)    
            
        elif winning_value >= 41 and winning_value <= 50: 
            jackpot = int(float(game.amount)*1)
            
        elif winning_value >= 51 and winning_value <= 60: 
            jackpot = int(float(game.amount)*1.5)       
    
        elif winning_value >= 61 and winning_value <= 63: 
            jackpot = int(float(game.amount)*2)   
            
        elif winning_value == 64: 
            jackpot = int(float(game.amount)*8)          
                
        player1.increase_chips(wb, ws, jackpot)
        wb.save('wallets.xlsx')
        time.sleep(2)
        if jackpot >= game.amount:
            emoji = '🍀 '
        else:
            emoji = ' ☘️ '
            
        badi_bot.send_message(game.chat_id,emoji+player1.firstname+' won ' +str(jackpot)+f' {currency}!')
        time.sleep(2)
        game.limit +=1
        if game.limit == 3:
            print('im here')
            button_slots = {"inline_keyboard" : [[{"text" :'🛠️ Repair', "callback_data" : 'repair_now'}]]}            
            button_slots = json.dumps(button_slots)
            badi_bot.edit_message_inline_id(game.inline_message_id,"The slots machine broke down",button_slots)
            game.lock = 1
        
    except:
        print('error')
        #badi_bot.send_message(game.chat_id,'❌ Error, please restart the game')
    return True 



def all_chips(first_chat_id):
    slots_amount = 0 #ws['M' + str(2)].value
    owners =''
    total = 0
    j=2
    for i in ws:
        if str(ws['A' + str(j)].value) == 'None' or ws['C' + str(j)].value == 0:
            j+=1
            continue
        owners += str(ws['A' + str(j)].value)+' own '+ str(ws['C' + str(j)].value)+f' {currency}.\n'
        if str(ws['C' + str(j)].value).isdigit():

            total += ws['C' + str(j)].value

        j+=1
    badi_bot.send_message(first_chat_id,owners+'\n________\nTotal: ' + str(total)+' + '+str(slots_amount) +' = '+ str(total+slots_amount))
    j=0

def show_wallet(current_update):
     user_id = current_update['message']['from']['id']
     firstname = current_update['message']['from']['first_name']
     chat_id = current_update['message']['chat']['id']

     
     player1 = Player(user_id, firstname)
     playercoins = player1.chips(wb, ws)
     message_text = f'{firstname} {currency} 𝘄𝗮𝗹𝗹𝗲𝘁 𝗯𝗮𝗹𝗮𝗻𝗰𝗲:\n\n{playercoins} ©️'
     badi_bot.send_message(chat_id,message_text)
     

def addchips(current_update):
    text = current_update['message']['text']
    text_lst = text.split(' ')
    amount = int(text_lst[2])
    user_id = int(text_lst[1])
    player1 = Player(user_id,'idc')
    player1.increase_chips(wb, ws, amount)
    badi_bot.send_message(user_id,f'✅ You received {amount} {currency}.') 
    badi_bot.send_message(droid_core,f'✅ You sent {amount} {currency}.') 
    
    
def start_dices_func(game):
    try:
        inline_msg_id = game.inline_message_id
        x = telethon.utils.resolve_inline_message_id(inline_msg_id)
        game.message_id = x[0] 
        game.chat_id = telethon.utils.get_peer_id(x[1],add_mark=True)
        print(game.chat_id)
        f = badi_bot.send_message(game.chat_id,'🎲 Game started!\nAnd may the odds be ever in your favor.')
        print(f)
        if str(f) != '<Response [200]>':
            badi_bot.send_message()
       # a_new_game = multiprocessing.Process(target=start_new_game,args=[game])
       # a_new_game.start()
        player1 = Player(game.player1_id,game.player1_firstname)
        player2 = Player(game.player2_id,game.player2_firstname)
        player1_chips = player1.chips(wb, ws)
        player2_chips = player2.chips(wb, ws)
        jackpot = 0
        if player1_chips < game.amount or player2_chips < game.amount:
            return False
        jackpot = game.amount*2
        player1.reduce_chips(wb, ws, game.amount)
        player2.reduce_chips(wb, ws, game.amount) 
        player1_wins = 0
        player2_wins = 0
        while player1_wins == player2_wins:
            badi_bot.send_message(game.chat_id,player1.firstname + ' roll: ')
            num_player1 = roll(game.chat_id)
            badi_bot.send_message(game.chat_id,player2.firstname + ' roll: ')
            num_player2 = roll(game.chat_id)
            
            if num_player1 > num_player2:
                player1_wins += 1
                winner = player1
    
            elif num_player2 > num_player1:
                player2_wins += 1
                winner = player2
                
        winner.increase_chips(wb, ws, jackpot)
        wb.save('wallets.xlsx')
        badi_bot.send_message(game.chat_id,'🍀 '+winner.firstname+' won ' +str(game.amount)+f' {currency}!')
        Games_Lst.remove(game)
    except:
        print('error')
        #badi_bot.send_message(game.chat_id,'❌ Error, please restart the game')
    return True
   
   
   


def awnser_to_inline_dice(inline_query_id):
    button_slots = {"inline_keyboard" : [[{"text" :'📍 Spin', "callback_data" : 'spin_now'}],[{"text" :f'1 {currency_symbol}', "callback_data" : 'add-1-chip'},{"text" :f'2 {currency_symbol}', "callback_data" : 'add-2-chip'},{"text" :f'4 {currency_symbol}', "callback_data" : 'add-4-chip'},{"text" :f'8 {currency_symbol}', "callback_data" : 'add-8-chip'},{"text" :f'16 {currency_symbol}', "callback_data" : 'add-16-chip'}]]} 
    button_dices = {"inline_keyboard" : [[{"text" :'🕹️ Join Match', "callback_data" : 'join_dices'},{"text" :'🎲 Start', "callback_data" : 'start_dices'}],[{"text" :f'1 {currency_symbol}', "callback_data" : 'add_1_chip'},{"text" :f'2 {currency_symbol}', "callback_data" : 'add_2_chip'},{"text" :f'4 {currency_symbol}', "callback_data" : 'add_4_chip'},{"text" :f'8 {currency_symbol}', "callback_data" : 'add_8_chip'},{"text" :f'16 {currency_symbol}', "callback_data" : 'add_16_chip'}]]} 
    dices = {'description':'start a new dices match','thumb_url':'https://cdn.pixabay.com/photo/2012/04/05/01/24/dice-25637_1280.png' , 'title': 'Dice Game', 'type': 'article', 'id': 'request_to_start_dices', 'input_message_content': {'message_text':f'Hey, I would like to play dices with you on \n0 {currency}.'},'reply_markup':button_dices}
    slots = {'description':'Slots gameplay','thumb_url':'https://cdn.pixabay.com/photo/2013/07/13/12/37/slot-machine-159972_1280.png' , 'title': 'Slots', 'type': 'article', 'id': 'request_to_start_slots', 'input_message_content': {'message_text':'🎰 If you’re looking for a thrill, then you better take chances on playing a 3-reel slot game!\n\n𝗠𝘂𝗹𝘁𝗶𝗽𝗹𝗶𝗲𝗿: 0'},'reply_markup':button_slots}  
    results1 = [dices,slots]     
    results1 = json.dumps(results1) 
    badi_bot.answer_inline_query(inline_query_id,results1) 


def pm(chat_id,text):
    badi_bot.send_message(int(chat_id),text)



def check_amount(amount,player1):
    #return int = 0 incase of error 
    #else retrun chips amount
    playerchips = player1.chips(wb, ws)
    numstest = amount.isdigit()
    if numstest == False:
        return 0
    amount = int(amount)
    if amount > playerchips or amount <= 0:
        return 0
    return amount
    
def amoun_test_max_min(min_chips,max_chips,player_amount):
    if player_amount <= max_chips and player_amount >= min_chips:
        return True
    return 0



def withdraw(user_id,chat_id,firstname,message_id):
    player1=Player(user_id,firstname)
    with_addy = player1.withdrawal_addy(wb, ws)
    text = f'Your current withdrawal {currency} address is:\n\n{with_addy}\n\n Tap Continue to select an amount to send.'
    button = {"inline_keyboard" : [[{"text" :'📝 Set new withdrawal address', "callback_data" : 'new_address'}],[{"text" :'⏩ Continue', "callback_data" : 'select_amount'}]]}        

    badi_bot.edit_message(chat_id,message_id,text)
    badi_bot.edit_inline(chat_id,message_id,button)
    

def bold(str1):
    str2 =''
    str3 = '𝟬𝟭𝟮𝟯𝟰𝟱𝟲𝟳𝟴𝟵'
    for i in str1:  
        str2 += str3[int(i)]
    return str2
            
     
def Lock_Wallets(wallets_lock): #2FIX
    
    while wallets_lock:
        time.sleep(1)
    wallets_lock = True    
    
def UnLock_Wallets(wallets_lock): #2FIX
    wallets_lock = False     

def wallet_button(current_update):
     if 'message' in current_update:
         user_id = current_update['message']['from']['id']
         firstname = current_update['message']['from']['first_name']
         chat_id = current_update['message']['chat']['id']
     else:
         user_id = current_update['callback_query']['from']['id']
         firstname = current_update['callback_query']['from']['first_name']
         chat_id = current_update['callback_query']['message']['chat']['id']
         
     
     player1 = Player(user_id, firstname)
     playercoins = player1.chips(wb, ws)
     message_text = f'𝗬𝗼𝘂𝗿 {currency} 𝘄𝗮𝗹𝗹𝗲𝘁 𝗯𝗮𝗹𝗮𝗻𝗰𝗲:\n\n{playercoins} ©️'
     
     button = {"inline_keyboard" : [[{"text" :'📥 Deposit', "callback_data" : 'deposit_chips'},{"text" :'📤 Withdraw', "callback_data" : 'withdraw_chips'}],[{"text" :'⛏️ Mine', "callback_data" : 'mine_chips'}]]}     
     badi_bot.send_inline_callback_button(chat_id,message_text,button)

def private_keyboard(current_update):
    chat_id = current_update['message']['chat']['id']
    text = current_update['message']['text']
    
    if text == '/start':
        keyboard = {"resize_keyboard" : True, 'one_time_keyboard' : False ,"keyboard" : [[{"text" : "💰 Wallet"},{"text" : "🎲 Play"}],[{"text" : "💡 Support"},{"text" : "🏷️ Free Market"}]]}
        keyboard = json.dumps(keyboard)
        badi_bot.open_keyboard(chat_id,'Successfully completed.',keyboard)

    

def roll(first_chat_id):
    r = badi_bot.send_dice(first_chat_id,'🎲')
    r_dict = r.json()
    while 'result' not in r_dict:
        time.sleep(1)
    time.sleep(4)     
    winning_value = r_dict['result']['dice']['value']
    winning_value = str(winning_value)
    if winning_value =='1':
        winning_num = '1️⃣'
    if winning_value =='2':
        winning_num = '2️⃣'
    if winning_value =='3':
        winning_num = '3️⃣'
    if winning_value =='4':
        winning_num = '4️⃣'
    if winning_value =='5':
        winning_num = '5️⃣'
    if winning_value =='6':
        winning_num = '6️⃣'
    badi_bot.send_message(first_chat_id,'💃 Number: ' +winning_num)
    time.sleep(4)
    return winning_num

def restart():
        print ("Restarting...")
        os.system("python telegramGames.py")
        time.sleep(10)
        return True
    
def main_menu_tggames():
    new_offset = 0
    first_update_id = 0
    timeout_restart = time.time()+ 3600 #Auto Restart
    timeout_spin = time.time()

    

    
    print('Leggo \n')
    while True:
        
      if time.time() > timeout_restart:
          timeout_restart = time.time() + 3600
          restart()

      all_updates = badi_bot.get_updates(new_offset)
      if len(all_updates) > 0:
            for current_update in all_updates:
                current_update_dumped = json.dumps(current_update)
                print('_____________________\n')
                pprint.pprint(json.loads(current_update_dumped))
                print('_____________________\n')
                first_update_id = current_update['update_id']
                
                if 'chosen_inline_result' in current_update:
                    if current_update['chosen_inline_result']['result_id'] == 'request_to_start_dices':
                         Games_Lst_item = NewGame(None,None,current_update['chosen_inline_result']['from']['id'],0, current_update['chosen_inline_result']['inline_message_id'], current_update['chosen_inline_result']['from']['first_name'],None,None)
                         Games_Lst.append(Games_Lst_item)
                         
                    if current_update['chosen_inline_result']['result_id'] == 'request_to_start_slots':
                         Games_Lst_item = NewSlotsGame(None,None,current_update['chosen_inline_result']['from']['id'],0, current_update['chosen_inline_result']['inline_message_id'], current_update['chosen_inline_result']['from']['first_name'])
                         Slots_Games_Lst.append(Games_Lst_item)
                         
                         
                         
                   
                
                if 'message' in current_update:
                    if 'chat' in current_update['message']:
                        chat_id = current_update['message']['chat']['id']
                        if current_update['message']['chat']['type'] == 'private':
                            if 'text' in current_update['message']:
                                private_keyboard(current_update)                                
                    if 'text' in current_update['message']:
                        if '/allchips' in current_update['message']['text'] and int(current_update['message']['from']['id']) == int(my_tg_id):
                            all_chips(chat_id)
                        
                        if '/wallet' in current_update['message']['text']:
                            show_wallet(current_update)
                        
                        if current_update['message']['text'] == "💡 Support":
                            badi_bot.force_reply(chat_id,f'🆘 In order to get support please check out our forum {groupforum}.\n if you still need help reply to this message.\n\n⬇️ Enter Support Ticket ⬇️')
                            
                        if current_update['message']['text'] == "🏷️ Free Market":
                            badi_bot.send_message(chat_id,f'🏪 Join our Free Market to spend {currency},\n{freemarket}')                       
                        
                        if current_update['message']['text'] == "💰 Wallet":
                            wallet_button(current_update)
                            
                        if current_update['message']['text'] == '🎲 Play':
                            badi_bot.send_inline_button(chat_id, '🎲',f'💬 Choose conversation to play on\n\n⚠️Make sure that {bot_name} found inside of the chat you would like to play on.','👁️‍🗨️ Select')
                        
                        if '/addchips' in current_update['message']['text'] and int(current_update['message']['from']['id']) == int(my_tg_id):
                            #FORMAT  /addchips userid amount
                            addchips(current_update)
                        
                        if '/pm' in current_update['message']['text'] and str(current_update['message']['from']['id']) == str(my_tg_id):
                            #FORMAT    /pm userid message
                            lst1 = (current_update['message']['text']).split(' ')
                            user_id_to_pm = lst1[1]
                            lst1.pop(0)
                            lst1.pop(0)
                            text_to_pm =''
                            for i in lst1:
                                text_to_pm += i
                                text_to_pm +=' '
                                
                            print(lst1)
                            pm(user_id_to_pm,text_to_pm)
                            
        
                    if 'reply_to_message' in current_update['message']:
                        try:
                            if current_update['message']['reply_to_message']['text'] == '✏️ Insert new address:':
                                new_addy = current_update['message']['text']
                                user_id = current_update['message']['chat']['id']
                                firstname = current_update['message']['chat']['first_name']
                                player1 = Player(user_id, firstname)
                                qtest = player1.set_withdrawal_addy(new_addy, wb, ws)
                                if qtest == True:
                                    badi_bot.send_message(chat_id,'✅ New address inserted successfully')
                                else:
                                    badi_bot.send_message(chat_id,'❌ Error')
                        except:
                                print("An exception occurred")
                         
                    if 'reply_to_message' in current_update['message']:
                        try:
                        
                            if 'text' in current_update['message']['reply_to_message']:

                                if current_update['message']['reply_to_message']['text'] == f'💳 Insert the amount of {currency} you wish to withdraw:':
                                  
                                    amount = current_update['message']['text']                                                     
                                    user_id = current_update['message']['chat']['id']
                                    firstname = current_update['message']['chat']['first_name']
                                    player_send = Player(user_id, firstname)
                                    #LOCK
                                    amount_test = check_amount(amount,player_send)
                                    player_recive_id = str(player_send.withdrawal_addy(wb, ws))[2:]
                                    if player_recive_id.isdigit():
                                        player_recive_id = int(player_recive_id)          
                                    player_recive = Player(player_recive_id, 'doesnt matter')
                                
    
                                    if amount_test != 0:
                                        max_chips = 500
                                        min_chips = 10
                                        amount_test_max_min = amoun_test_max_min(min_chips,max_chips,amount_test)
                                        if current_update['message']['reply_to_message']['text'] == f'💳 Insert the amount of {currency} you wish to withdraw:':
                                            min_chips = 1
                                            amount_test_max_min = amoun_test_max_min(min_chips,max_chips,amount_test)
                                        
                                        
                                        if amount_test_max_min == True:
                                            player_send.reduce_chips(wb, ws, amount_test)
                                            player_recive.increase_chips(wb,ws,amount_test)
                                            badi_bot.send_message(chat_id, f'✅ {amount_test} {currency} sent successfully')
                                            badi_bot.send_message(player_recive.user_id, f'✅ {amount_test} {currency} recived successfully from {firstname}')
                                           
                                                
                                        else:
                                            badi_bot.send_message(chat_id,f'❌ Error please select higher amount than {min_chips}')
                                    else:
                                        badi_bot.send_message(chat_id,'❌ Error')        
                                    #UNLOCK
                        except:
                                print("An exception occurred")       

                
                if 'callback_query' in current_update:
                     user_id = current_update['callback_query']['from']['id']
                     firstname = current_update['callback_query']['from']['first_name']
                     callback = current_update['callback_query']['data']
                     call_back_id = current_update['callback_query']['id']        
                     
                     if 'add_' in callback:
                        for game in Games_Lst:
                            if game.player1_id == user_id and game.inline_message_id == current_update['callback_query']['inline_message_id']:
                                 x = callback.split('_')  
                                 x = int(x[1])    
                                 game.change_chips(x,call_back_id)
                                 print(game.amount)
                                 
                            elif game.player1_id != user_id and game.inline_message_id == current_update['callback_query']['inline_message_id']:
                                     badi_bot.awnser_call_back_alert(call_back_id,'❌ You are not the creator of this game.')
                     
                     if 'add-' in callback:
                        for game in Slots_Games_Lst:
                            if game.player1_id == user_id and game.inline_message_id == current_update['callback_query']['inline_message_id']:
                                 x = callback.split('-')  
                                 x = int(x[1])    
                                 game.change_chips(x,call_back_id)
                                 print(game.amount)
                                 
                            elif game.player1_id != user_id and game.inline_message_id == current_update['callback_query']['inline_message_id']:
                                     badi_bot.awnser_call_back_alert(call_back_id,'❌ You are not the creator of this game.')
                                                     
                    
                    
                  
                     if callback == 'join_dices':
                        player2 = Player(user_id,firstname)
                        player2_chips = player2.chips(wb, ws)
                        for game in Games_Lst:
                            if player2.user_id != game.player1_id and game.inline_message_id == current_update['callback_query']['inline_message_id'] and player2_chips >= game.amount and game.amount>0:
                                game.player2_id = player2.user_id
                                game.player2_firstname = player2.firstname
                                button_dices = {"resize_inline_keyboard": True ,"inline_keyboard" : [[{"text" :'🕹️ Join Match', "callback_data" : 'join_dices'},{"text" :'🎲 Start', "callback_data" : 'start_dices'}],[{"text" :f'1 {currency_symbol}', "callback_data" : 'add_1_chip'},{"text" :f'2 {currency_symbol}', "callback_data" : 'add_2_chip'},{"text" :f'4 {currency_symbol}', "callback_data" : 'add_4_chip'},{"text" :f'8 {currency_symbol}', "callback_data" : 'add_8_chip'},{"text" :f'16 {currency_symbol}', "callback_data" : 'add_16_chip'}]]} 
                                button_dices = json.dumps(button_dices)     
                                badi_bot.edit_message_inline_id(game.inline_message_id,f'Hey, I would like to play dices with you on \n{game.amount} {currency}.\n{game.player1_firstname} vs {game.player2_firstname}',button_dices)
                            elif player2_chips < game.amount and player2.user_id != game.player1_id and game.inline_message_id == current_update['callback_query']['inline_message_id']:
                                badi_bot.awnser_call_back_alert(call_back_id,'❌ Unfortunately you do not have enough {currency} to complete this operation,\nMake sure to mine {currency} through {bot_username} before trying again.')
                            elif player2.user_id == game.player1_id and game.inline_message_id == current_update['callback_query']['inline_message_id'] and player2_chips >= game.amount:
                                  badi_bot.awnser_call_back_alert(call_back_id,"❌ You can't play with yourself")
                                  
                            elif player2.user_id != game.player1_id and game.inline_message_id == current_update['callback_query']['inline_message_id'] and player2_chips >= game.amount and game.amount==0:
                                badi_bot.awnser_call_back_alert(call_back_id,"❌ You can't join none {currency} game")    
                            
                            
                     if callback == 'start_dices':
                         for game in Games_Lst:
                             if game.player1_id == user_id and game.player2_id != None and game.inline_message_id == current_update['callback_query']['inline_message_id']:
                                 try:
                                     chat_id = game.chat_id
                                     start_dices_func(game)
                                 except:
                                     print("An exception occurred")
                                     badi_bot.awnser_call_back_alert(call_back_id,f'❌ Failed!\nPlease add {bot_username} to this chat before trying again.\nThanks!')
                             elif game.player1_id != user_id and game.player2_id != None and game.inline_message_id == current_update['callback_query']['inline_message_id']:
                                  badi_bot.awnser_call_back_alert(call_back_id,'❌ You are not the creator of this game.')
                             elif game.player2_id == None and game.player1_id == user_id and game.inline_message_id == current_update['callback_query']['inline_message_id'] :
                                  badi_bot.awnser_call_back_alert(call_back_id,"❌ Please be patient and wait for another player to join this game.")
                      
                                  
                     if callback == 'repair_now':
                        for game in Slots_Games_Lst:
                            if game.inline_message_id == current_update['callback_query']['inline_message_id']:
                                button_slots = {"inline_keyboard" : [[{"text" :'📍 Spin', "callback_data" : 'spin_now'}],[{"text" :'1 {currency_symbol}', "callback_data" : 'add-1-chip'},{"text" :'2 {currency_symbol}', "callback_data" : 'add-2-chip'},{"text" :'4 {currency_symbol}', "callback_data" : 'add-4-chip'},{"text" :'8 {currency_symbol}', "callback_data" : 'add-8-chip'},{"text" :'16 {currency_symbol}', "callback_data" : 'add-16-chip'}]]} 
                                button_slots = json.dumps(button_slots)                    
                                badi_bot.edit_message_inline_id(game.inline_message_id,f"🎰 If you’re looking for a thrill, then you better take chances on playing a 3-reel slot game!\n\n𝗠𝘂𝗹𝘁𝗶𝗽𝗹𝗶𝗲𝗿: {game.amount}",button_slots)
                                game.lock = 0
                                    
                     if callback == 'spin_now':
                         if  time.time() > timeout_spin:
                             min_spin = 5
                             player1 = Player(user_id,firstname)
                             player_chips = player1.chips(wb, ws)
                             for game in Slots_Games_Lst:
                                 if game.inline_message_id == current_update['callback_query']['inline_message_id'] and player_chips >= game.amount and game.amount >=min_spin and game.lock ==0:
                                     try:
                                         game.player1_id = user_id
                                         game.player1_firstname = firstname
                                         chat_id = game.chat_id
                                         spin_done = start_slots_func(game)
                                         if spin_done == True:
                                             timeout_spin = time.time() + 15
                                     except:
                                         print("An exception occurred")
                                         badi_bot.awnser_call_back_alert(call_back_id,f'❌ Failed!\nPlease add {bot_username} to this chat before trying again.\nThanks!')
    
                                 elif game.inline_message_id == current_update['callback_query']['inline_message_id'] and player_chips >= game.amount and game.amount < min_spin:
                                     badi_bot.awnser_call_back_alert(call_back_id,f"❌ You can't spin on less than {min_spin} {currency}")                        
                            
                         elif time.time() < timeout_spin:
                             badi_bot.awnser_call_back_alert_black(call_back_id,'⌛ Cooldown Remaining Time: '+str(int(timeout_spin-time.time()))+' seconds.')
                                
                                   
                     if 'message' in current_update['callback_query']:
                       chat_id = current_update['callback_query']['message']['chat']['id']
                       message_id = current_update['callback_query']['message']['message_id']
                     
                        
                     
                        
                     if current_update['callback_query']['data'] == 'deposit_chips':
                         badi_bot.send_message(chat_id,f'Your personal {currency} address:')   
                         player1 = Player(user_id,firstname)
                         player_coins = 'DC'+ str(player1.addy(wb, ws))
                         badi_bot.send_message(chat_id,player_coins)  
                         
                     elif current_update['callback_query']['data'] == 'withdraw_chips':    
                         withdraw(user_id,chat_id,firstname,message_id)
                         
                         
                     if current_update['callback_query']['data'] == 'new_address':    
                         badi_bot.force_reply(chat_id,'✏️ Insert new address:')
                       
                     if current_update['callback_query']['data'] == 'select_amount':    
                         badi_bot.force_reply(chat_id,f'💳 Insert the amount of {currency} you wish to withdraw:')
                         
                
                     if current_update['callback_query']['data'] == 'mine_chips':
                         player1 = Player(user_id,firstname)
                         timeout_mine = player1.check_mine_time()
                         if  time.time() > timeout_mine:
                             player1.mine(time.time())
    
                         elif time.time() < timeout_mine:
                            timeleft = int(timeout_mine-time.time())
                            hoursleft = timeleft // 3600
                            minsleft = timeleft // 60
                            secleft = timeleft%60
                            badi_bot.awnser_call_back_alert_black(call_back_id,'⌛ Cooldown Remaining Time: '+str(hoursleft)+':'+str(minsleft)+':'+str(secleft))
                                

                         
                         
                         
                      
                if 'inline_query' in current_update:
                    if current_update['inline_query']['query'] == '🎲':
                        inline_q_id = current_update['inline_query']['id']
                        awnser_to_inline_dice(inline_q_id)
                         
                        
                new_offset = first_update_id + 1




if __name__ == '__main__':
    try:
        main_menu_tggames()


    except KeyboardInterrupt:
       print('\nrestart \  stop')     
       
       x = input('>')
       if x == 'restart':
           restart()

           exit()
       elif x =='stop':
           exit()

       else:
           print('failed')






